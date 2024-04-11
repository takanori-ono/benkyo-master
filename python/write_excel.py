from openpyxl import Workbook, load_workbook, utils
from openpyxl.worksheet.table import Table, TableStyleInfo

# 新しいワークブックを作成
# wb = Workbook()
# テンプレートを読み込んでみる
wb = load_workbook('template.xlsx')

# デフォルトのアクティブなワークシートを取得
ws = wb.active

# セルに値を設定
#ws['A1'] = 'Hello'
ws.cell(1,1).value = 'Hello'
ws.cell(1,2).value = 'world'

ws.cell(2,1).value = 123
ws.cell(2,2).value = 456
ws.cell(3,1).value = 1000
ws.cell(3,2).value = 2000

# Tableを作成
table = Table(displayName='Table1', ref='A1:B3')

# テーブルのスタイルを設定
table_style = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
table.tableStyleInfo = table_style

ws.add_table(table)

ws.column_dimensions[utils.get_column_letter(1)].width = 30
ws.column_dimensions[utils.get_column_letter(2)].width = 20
# ws.row_dimensions[1].height = 15
# ws.row_dimensions[2].height = 15
# ws.row_dimensions[3].height = 15
#ws.row_dimension

# Excelファイルを保存
wb.save('out_example.xlsx')
